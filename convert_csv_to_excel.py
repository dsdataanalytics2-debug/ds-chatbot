#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
📊 CSV to Excel Converter
CSV ফাইলগুলো Excel ফরম্যাটে রূপান্তর করে
"""

import csv
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

def csv_to_excel(csv_file, excel_file):
    """CSV ফাইল Excel এ রূপান্তর করে"""
    try:
        # Read CSV file
        with open(csv_file, 'r', encoding='utf-8') as file:
            csv_reader = csv.reader(file)
            data = list(csv_reader)
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        
        # Add data to Excel
        for row_idx, row in enumerate(data, 1):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                
                # Style header row
                if row_idx == 1:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center")
                else:
                    cell.alignment = Alignment(horizontal="left")
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save Excel file
        wb.save(excel_file)
        print(f"✅ {csv_file} সফলভাবে {excel_file} এ রূপান্তর হয়েছে!")
        
        return True
        
    except Exception as e:
        print(f"❌ {csv_file} রূপান্তর করতে সমস্যা: {str(e)}")
        return False

def main():
    print("🏥 DIGITAL SEBE CHATBOT - CSV to Excel Converter")
    print("=" * 60)
    
    # Convert medicine data
    if csv_to_excel('medicine_data.csv', 'medicine_data.xlsx'):
        print("💊 মেডিকেল ডেটা Excel ফাইলে সেভ হয়েছে")
    
    print()
    
    # Convert phone numbers
    if csv_to_excel('sample_phone_numbers.csv', 'sample_phone_numbers.xlsx'):
        print("📱 ফোন নম্বর Excel ফাইলে সেভ হয়েছে")
    
    print()
    print("🎉 সব ফাইল সফলভাবে রূপান্তর হয়েছে!")
    print("📁 Excel ফাইলগুলো: medicine_data.xlsx, sample_phone_numbers.xlsx")

if __name__ == "__main__":
    main()
